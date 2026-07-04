import json
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    messagebox.showerror("缺少库", "请先运行: pip install openpyxl")
    sys.exit(1)

CONFIG_FILE = "update_config.json"

def save_config(data):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return {
            "data_source_path": "",
            "template_path": "",
            "output_suffix": "_已更新",
            "data_mappings": [],
            "image_mappings": []
        }
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel报告一键更新器")
        self.root.geometry("950x720")
        self.config = load_config()

        # ---------- 基础路径 ----------
        frm_path = ttk.LabelFrame(root, text="基础路径")
        frm_path.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(frm_path, text="数据源路径:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.src_path_var = tk.StringVar(value=self.config.get("data_source_path", ""))
        ttk.Entry(frm_path, textvariable=self.src_path_var, width=80).grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(frm_path, text="浏览", command=self.browse_src).grid(row=0, column=2, padx=5)

        ttk.Label(frm_path, text="模板路径:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        self.tpl_path_var = tk.StringVar(value=self.config.get("template_path", ""))
        ttk.Entry(frm_path, textvariable=self.tpl_path_var, width=80).grid(row=1, column=1, padx=5, pady=2)
        ttk.Button(frm_path, text="浏览", command=self.browse_tpl).grid(row=1, column=2, padx=5)

        ttk.Label(frm_path, text="输出文件后缀:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.suffix_var = tk.StringVar(value=self.config.get("output_suffix", "_已更新"))
        ttk.Entry(frm_path, textvariable=self.suffix_var, width=20).grid(row=2, column=1, sticky=tk.W, padx=5)

        # ---------- 映射管理 ----------
        nb = ttk.Notebook(root)
        nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # ---- 数据映射页 ----
        frm_data = ttk.Frame(nb)
        nb.add(frm_data, text="数据映射")
        self.data_tree = ttk.Treeview(frm_data, columns=("source", "target"), show="headings", height=8)
        self.data_tree.heading("source", text="数据源单元格 (如 Sheet1!B2)")
        self.data_tree.heading("target", text="模板目标单元格 (如 Sheet1!D5)")
        self.data_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        btn_frm_data = ttk.Frame(frm_data)
        btn_frm_data.pack(fill=tk.X, padx=5, pady=2)
        ttk.Button(btn_frm_data, text="添加", command=self.add_data_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frm_data, text="编辑选中", command=self.edit_data_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frm_data, text="删除选中", command=lambda: self.delete_selected(self.data_tree, "data")).pack(side=tk.LEFT, padx=5)

        # ---- 图片映射页 ----
        frm_img = ttk.Frame(nb)
        nb.add(frm_img, text="图片映射")
        self.img_tree = ttk.Treeview(frm_img, columns=("number", "folder", "target", "width", "height"), show="headings", height=8)
        self.img_tree.heading("number", text="图片编号")
        self.img_tree.heading("folder", text="图片文件夹")
        self.img_tree.heading("target", text="目标单元格")
        self.img_tree.heading("width", text="宽度(cm)")
        self.img_tree.heading("height", text="高度(cm)")
        self.img_tree.column("number", width=100)
        self.img_tree.column("folder", width=250)
        self.img_tree.column("target", width=120)
        self.img_tree.column("width", width=80)
        self.img_tree.column("height", width=80)
        self.img_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        btn_frm_img = ttk.Frame(frm_img)
        btn_frm_img.pack(fill=tk.X, padx=5, pady=2)
        ttk.Button(btn_frm_img, text="添加", command=self.add_image_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frm_img, text="编辑选中", command=self.edit_image_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frm_img, text="删除选中", command=lambda: self.delete_selected(self.img_tree, "image")).pack(side=tk.LEFT, padx=5)

        ttk.Button(root, text="一键更新报告", command=self.run_update).pack(pady=10)

        self.refresh_data_tree()
        self.refresh_image_tree()

    # ---------- 路径浏览 ----------
    def browse_src(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.src_path_var.set(path)

    def browse_tpl(self):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.tpl_path_var.set(path)

    # ---------- 数据映射操作 ----------
    def add_data_mapping(self):
        self._data_dialog(None)

    def edit_data_mapping(self):
        selected = self.data_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "请先选择一条映射")
            return
        idx = self.data_tree.index(selected[0])
        item = self.config["data_mappings"][idx]
        self._data_dialog(idx, item)

    def _data_dialog(self, edit_idx, item=None):
        popup = tk.Toplevel(self.root)
        popup.title("编辑数据映射" if item else "添加数据映射")
        ttk.Label(popup, text="源单元格 (如 Sheet1!B2):").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        src_entry = ttk.Entry(popup, width=35)
        src_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(popup, text="目标单元格 (如 Sheet1!D5):").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        tgt_entry = ttk.Entry(popup, width=35)
        tgt_entry.grid(row=1, column=1, padx=5, pady=5)
        if item:
            src_entry.insert(0, item["source_cell"])
            tgt_entry.insert(0, item["target_cell"])

        def save():
            src = src_entry.get().strip()
            tgt = tgt_entry.get().strip()
            if not src or not tgt:
                messagebox.showwarning("输入不完整", "两个单元格都不能为空")
                return
            new_map = {"source_cell": src, "target_cell": tgt}
            if edit_idx is not None:
                self.config["data_mappings"][edit_idx] = new_map
            else:
                self.config.setdefault("data_mappings", []).append(new_map)
            self.refresh_data_tree()
            popup.destroy()
        ttk.Button(popup, text="确定", command=save).grid(row=2, column=0, columnspan=2, pady=10)

    def refresh_data_tree(self):
        for i in self.data_tree.get_children():
            self.data_tree.delete(i)
        for m in self.config.get("data_mappings", []):
            self.data_tree.insert("", tk.END, values=(m["source_cell"], m["target_cell"]))

    # ---------- 图片映射操作 ----------
    def add_image_mapping(self):
        self._image_dialog(None)

    def edit_image_mapping(self):
        selected = self.img_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "请先选择一条映射")
            return
        idx = self.img_tree.index(selected[0])
        item = self.config["image_mappings"][idx]
        self._image_dialog(idx, item)

    def _image_dialog(self, edit_idx, item=None):
        popup = tk.Toplevel(self.root)
        popup.title("编辑图片映射" if item else "添加图片映射")
        ttk.Label(popup, text="图片编号 (用于匹配文件名):").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        num_entry = ttk.Entry(popup, width=30)
        num_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(popup, text="图片文件夹:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        folder_var = tk.StringVar()
        folder_entry = ttk.Entry(popup, textvariable=folder_var, width=30)
        folder_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(popup, text="浏览", command=lambda: folder_var.set(filedialog.askdirectory())).grid(row=1, column=2, padx=5)
        ttk.Label(popup, text="目标单元格 (如 Sheet1!B10):").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        tgt_entry = ttk.Entry(popup, width=30)
        tgt_entry.grid(row=2, column=1, padx=5, pady=5)
        ttk.Label(popup, text="宽度 (cm):").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        w_entry = ttk.Entry(popup, width=10)
        w_entry.grid(row=3, column=1, sticky=tk.W, padx=5)
        ttk.Label(popup, text="高度 (cm):").grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)
        h_entry = ttk.Entry(popup, width=10)
        h_entry.grid(row=4, column=1, sticky=tk.W, padx=5)

        if item:
            num_entry.insert(0, item["image_number"])
            folder_var.set(item["image_folder"])
            tgt_entry.insert(0, item["target_cell"])
            w_entry.insert(0, str(item["width_cm"]))
            h_entry.insert(0, str(item["height_cm"]))
        else:
            w_entry.insert(0, "3.5")
            h_entry.insert(0, "2.8")

        def save():
            num = num_entry.get().strip()
            folder = folder_var.get().strip()
            tgt = tgt_entry.get().strip()
            try:
                w = float(w_entry.get().strip())
                h = float(h_entry.get().strip())
            except ValueError:
                messagebox.showwarning("输入错误", "宽度和高度必须为数字")
                return
            if not num or not folder or not tgt:
                messagebox.showwarning("输入不完整", "所有字段必填")
                return
            new_map = {
                "image_number": num,
                "image_folder": folder,
                "target_cell": tgt,
                "width_cm": w,
                "height_cm": h
            }
            if edit_idx is not None:
                self.config["image_mappings"][edit_idx] = new_map
            else:
                self.config.setdefault("image_mappings", []).append(new_map)
            self.refresh_image_tree()
            popup.destroy()
        ttk.Button(popup, text="确定", command=save).grid(row=5, column=0, columnspan=3, pady=10)

    def refresh_image_tree(self):
        for i in self.img_tree.get_children():
            self.img_tree.delete(i)
        for m in self.config.get("image_mappings", []):
            self.img_tree.insert("", tk.END, values=(
                m["image_number"],
                m["image_folder"],
                m["target_cell"],
                m["width_cm"],
                m["height_cm"]
            ))

    def delete_selected(self, tree, map_type):
        selected = tree.selection()
        if not selected:
            return
        idx = tree.index(selected[0])
        if map_type == "data":
            del self.config["data_mappings"][idx]
            self.refresh_data_tree()
        else:
            del self.config["image_mappings"][idx]
            self.refresh_image_tree()

    def save_current_config(self):
        self.config["data_source_path"] = self.src_path_var.get()
        self.config["template_path"] = self.tpl_path_var.get()
        self.config["output_suffix"] = self.suffix_var.get()
        save_config(self.config)

    # ---------- 核心执行（增强调试） ----------
    def run_update(self):
        self.save_current_config()
        cfg = self.config

        if not cfg.get("data_source_path") or not os.path.exists(cfg["data_source_path"]):
            messagebox.showerror("错误", "数据源文件不存在")
            return
        if not cfg.get("template_path") or not os.path.exists(cfg["template_path"]):
            messagebox.showerror("错误", "模板文件不存在")
            return

        try:
            wb_src = load_workbook(cfg["data_source_path"], data_only=True)
        except Exception as e:
            messagebox.showerror("打开数据源失败", str(e))
            return
        try:
            wb = load_workbook(cfg["template_path"])
        except Exception as e:
            messagebox.showerror("打开模板失败", str(e))
            wb_src.close()
            return

        # ----- 数据写入 -----
        for i, m in enumerate(cfg.get("data_mappings", [])):
            try:
                if "!" not in m["source_cell"]:
                    raise ValueError("缺少 '!' 分隔符")
                if "!" not in m["target_cell"]:
                    raise ValueError("缺少 '!' 分隔符")
                src_sh, src_cell = m["source_cell"].split("!", 1)
                tgt_sh, tgt_cell = m["target_cell"].split("!", 1)

                if src_sh not in wb_src.sheetnames:
                    raise KeyError(f"数据源中不存在工作表：'{src_sh}'\n可用工作表：{wb_src.sheetnames}")
                if tgt_sh not in wb.sheetnames:
                    raise KeyError(f"模板中不存在工作表：'{tgt_sh}'\n可用工作表：{wb.sheetnames}")

                ws_src = wb_src[src_sh]
                ws_tgt = wb[tgt_sh]
                ws_tgt[tgt_cell].value = ws_src[src_cell].value
            except Exception as e:
                wb_src.close()
                wb.close()
                messagebox.showerror("数据写入出错",
                    f"映射 {i+1}:\n源 {m['source_cell']} → 目标 {m['target_cell']}\n错误：{e}")
                return

        # ----- 图片插入（带详细反馈） -----
        inserted_count = 0
        skipped_details = []  # 记录跳过的原因
        for i, m in enumerate(cfg.get("image_mappings", [])):
            try:
                number = m["image_number"]
                folder = Path(m["image_folder"])

                # 检查文件夹是否存在且确实为文件夹
                if not folder.exists():
                    skipped_details.append(f"映射{i+1}: 文件夹不存在 {folder}")
                    continue
                if not folder.is_dir():
                    skipped_details.append(f"映射{i+1}: 路径不是文件夹 {folder}")
                    continue

                # 列出文件夹内所有文件名（用于匹配和调试）
                folder_files = {}
                try:
                    for f in folder.iterdir():
                        if f.is_file():
                            folder_files[f.name.lower()] = f.name
                except Exception as e:
                    skipped_details.append(f"映射{i+1}: 无法读取文件夹 {folder}，错误: {e}")
                    continue

                if not folder_files:
                    skipped_details.append(f"映射{i+1}: 文件夹为空 {folder}")
                    continue

                # 查找图片（忽略扩展名大小写）
                img_path = None
                matched_name = None
                for ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif"]:
                    candidate_name = f"{number}{ext}".lower()
                    if candidate_name in folder_files:
                        matched_name = folder_files[candidate_name]
                        img_path = str(folder / matched_name)
                        break

                if img_path is None:
                    file_list = "\n".join(sorted(folder_files.values())[:15])
                    skipped_details.append(f"映射{i+1}: 未找到编号 '{number}' 的图片\n文件夹内容(前15个):\n{file_list}")
                    continue

                # 目标单元格检查
                if "!" not in m["target_cell"]:
                    skipped_details.append(f"映射{i+1}: 目标单元格格式错误 (缺少'!')")
                    continue
                tgt_sh, tgt_cell = m["target_cell"].split("!", 1)
                if tgt_sh not in wb.sheetnames:
                    skipped_details.append(f"映射{i+1}: 模板中不存在工作表 '{tgt_sh}'")
                    continue

                # 插入图片
                ws_tgt = wb[tgt_sh]
                img = XLImage(img_path)
                img.width = m["width_cm"]
                img.height = m["height_cm"]
                ws_tgt.add_image(img, tgt_cell)
                inserted_count += 1

            except Exception as e:
                wb_src.close()
                wb.close()
                messagebox.showerror("图片插入出错",
                    f"图片映射 {i+1}:\n编号 {m['image_number']}，目标 {m['target_cell']}\n错误：{e}")
                return

        # 保存前关闭数据源
        wb_src.close()

        # 保存新文件
        tpl_path = Path(cfg["template_path"])
        suffix = cfg.get("output_suffix", "_已更新")
        out_path = tpl_path.parent / f"{tpl_path.stem}{suffix}.xlsx"
        try:
            wb.save(str(out_path))
        except Exception as e:
            wb.close()
            messagebox.showerror("保存失败", str(e))
            return

        wb.close()

        # 调试反馈
        summary = f"数据映射: {len(cfg.get('data_mappings', []))} 条\n"
        summary += f"图片映射: {len(cfg.get('image_mappings', []))} 条\n"
        summary += f"成功插入图片: {inserted_count} 张\n"
        if skipped_details:
            summary += "\n未插入图片原因:\n" + "\n\n".join(skipped_details)
        else:
            summary += "\n所有图片均已插入。"

        messagebox.showinfo("执行结果", summary)

        # 自动打开
        try:
            os.startfile(out_path)
        except Exception:
            pass
        messagebox.showinfo("文件位置", f"报告已生成:\n{out_path}")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
